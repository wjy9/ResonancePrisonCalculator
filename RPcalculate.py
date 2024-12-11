import sys
import openpyxl as xl
import click

manu_dict = {}
ingr_dict = {}
effi_dict = {}
purch_dict = {}
price_dict = {}

class Recipe:
    def __init__(self):
        self.num = 0
        self.time = 0
        self.electric = 0
        self.ingredient = {}
        self.exp = 0.0
        self.order_price = 0.0

class Ingredient:
    def __init__(self):
        self.time = 0
        self.electric = 0
        self.money = 0

class Efficiency:
    def __init__(self):
        self.manu_time = 0
        self.base_material_cost = {}
        self.electric = 0
        self.money_cost = 0
        self.per_base_mat = {}
        self.per_electric = 0
        self.per_time = 0
        self.per_money = 0

class Purchase:
    def __init__(self):
        self.material = None
        self.ratio = 0

def read_table(filename):
    wb = xl.load_workbook(filename=filename, read_only=True)

    m_sheet = wb["生产车间"]
    i_sheet = wb["原料车间"]
    p_sheet = wb["原料采购"]
    o_sheet = wb["订单"]
    pr_sheet = wb["交易所"]

    for row in m_sheet.iter_rows(min_row=2):
        name = row[0].value
        mat = row[1].value
        num = row[2].value
        if name not in manu_dict:
            manu_dict[name] = Recipe()
        if (mat == "产量"):
            manu_dict[name].num = num
        elif (mat == "时间"):
            manu_dict[name].time = num
        elif (mat == "电量"):
            manu_dict[name].electric = num
        else:
            manu_dict[name].ingredient[mat] = num

    for row in i_sheet.iter_rows(min_row=2):
        name = row[0].value
        time = row[1].value
        num = row[2].value
        electric = row[3].value
        money = row[4].value if row[4].value else 0
        if name not in ingr_dict:
            ingr_dict[name] = Ingredient()
        ingr_dict[name].time = time / num
        ingr_dict[name].electric = electric * time / num
        ingr_dict[name].money = money * time / num

    for row in o_sheet.iter_rows(min_row=2):
        name = row[0].value
        exp = row[1].value
        money = row[2].value
        if name not in manu_dict:
            print("订单{}不在生产车间制造表中！".format(name))
            continue
        manu_dict[name].exp = exp / 60
        manu_dict[name].order_price = money / 60

    for row in p_sheet.iter_rows(min_row=2):
        name = row[0].value
        material = row[1].value
        ratio = row[2].value
        purch_dict[name] = Purchase()
        purch_dict[name].material = material
        purch_dict[name].ratio = ratio

    for row in pr_sheet.iter_rows(min_row=2):
        name = row[0].value
        price = row[1].value
        price_dict[name] = price

def expand_purchase(name):
    recipe = manu_dict[name]
    for mat in recipe.ingredient:
        if mat in purch_dict:
            new_name = name + "采购" + mat
            manu_dict[new_name] = Recipe()
            manu_dict[new_name].num = recipe.num
            manu_dict[new_name].time = recipe.time
            manu_dict[new_name].electric = recipe.electric
            manu_dict[new_name].ingredient = {}
            for m in recipe.ingredient:
                if m == mat:
                    new_m = purch_dict[mat].material
                    manu_dict[new_name].ingredient[new_m] = recipe.ingredient[m] / purch_dict[mat].ratio
                else:
                    manu_dict[new_name].ingredient[m] = recipe.ingredient[m]
            manu_dict[new_name].exp = recipe.exp
            manu_dict[new_name].order_price = recipe.order_price
            expand_purchase(new_name)

def calculate_efficiency(name):
    if name in effi_dict:
        return
    effi_dict[name] = Efficiency()
    effi_dict[name].manu_time = manu_dict[name].time / manu_dict[name].num
    effi_dict[name].electric = manu_dict[name].electric / manu_dict[name].num
    base_material_cost = {}
    for mat in manu_dict[name].ingredient:
        if mat in ingr_dict:
            # is a base material
            effi_dict[name].electric += ingr_dict[mat].electric * manu_dict[name].ingredient[mat] / manu_dict[name].num / manu_dict[name].time
            effi_dict[name].money_cost += ingr_dict[mat].money * manu_dict[name].ingredient[mat] / manu_dict[name].num
            if mat not in base_material_cost:
                base_material_cost[mat] = 0
            base_material_cost[mat] += manu_dict[name].ingredient[mat] / manu_dict[name].num
        elif mat in manu_dict:
            # is a product
            calculate_efficiency(mat)
            effi_dict[name].manu_time += effi_dict[mat].manu_time * manu_dict[name].ingredient[mat] / manu_dict[name].num
            effi_dict[name].electric += effi_dict[mat].electric * manu_dict[name].ingredient[mat] / manu_dict[name].num
            effi_dict[name].money_cost += effi_dict[mat].money_cost * manu_dict[name].ingredient[mat] / manu_dict[name].num
            for base_mat in effi_dict[mat].base_material_cost:
                if base_mat not in base_material_cost:
                    base_material_cost[base_mat] = 0
                base_material_cost[base_mat] += effi_dict[mat].base_material_cost[base_mat] * manu_dict[name].ingredient[mat] / manu_dict[name].num
        else:
            # buyable in store
            if mat in price_dict:
                effi_dict[name].money_cost += price_dict[mat] * manu_dict[name].ingredient[mat] / manu_dict[name].num
            else:
                print("error: {} no price".format(mat))
    effi_dict[name].base_material_cost = base_material_cost
    # for ingr in effi_dict[name].base_material_cost:
    #     exp_per_ingr = manu_dict[name].exp / effi_dict[name].base_material_cost[ingr]
    #     effi_dict[name].exp_per_base[ingr] = exp_per_ingr
    # exp_per_ele = manu_dict[name].exp / effi_dict[name].electric
    # effi_dict[name].exp_per_electric = exp_per_ele
    # exp_per_time = manu_dict[name].exp / effi_dict[name].manu_time
    # effi_dict[name].exp_per_time = exp_per_time
    # effi_dict[name].exp_per_money = manu_dict[name].exp / effi_dict[name].money_cost * 1000 if effi_dict[name].money_cost != 0 else 0

def calulate_target_per_resource(target):
    for product in effi_dict:
        value = 0
        if target == 1:
            value = manu_dict[product].exp
        else:
            value = manu_dict[product].order_price
        effi_dict[product].per_electric = value / effi_dict[product].electric
        effi_dict[product].per_time = value / effi_dict[product].manu_time
        effi_dict[product].per_money = value / effi_dict[product].money_cost * 1000 if effi_dict[product].money_cost != 0 else 0
        for ingr in effi_dict[product].base_material_cost:
            effi_dict[product].per_base_mat[ingr] = value / effi_dict[product].base_material_cost[ingr]

def calculate(enable_purchase):
    if enable_purchase:
        keys = list(manu_dict.keys())
        for product in keys:
            expand_purchase(product)
    for product in manu_dict:
        if manu_dict[product].exp > 0:
            calculate_efficiency(product)

@click.command
@click.option('--target', prompt="计算效率的目标 1.经验 2.订单利润", default=1)
@click.option('--sort_key', prompt="请选择排序依据，电量、时间、钱或某一种原料", default="电量")
@click.option('--output_num', prompt="显示最高N位，0为全显示", default=0)
def output(target, sort_key, output_num):
    calulate_target_per_resource(target)
    effi_list = effi_dict.items()
    if sort_key == "电量":
        effi_list = sorted(effi_list, key=lambda x : x[1].per_electric, reverse=True)
    elif sort_key == "时间":
        effi_list = sorted(effi_list, key=lambda x : x[1].per_time, reverse=True)
    elif sort_key == "钱":
        effi_list = sorted(effi_list, key=lambda x : x[1].per_money, reverse=True)
    else:
        effi_list = sorted(effi_list, key=lambda x : x[1].per_base_mat[sort_key] if sort_key in x[1].per_base_mat else 0, reverse=True)
    for idx, (k, v) in enumerate(effi_list):
        if output_num and idx >= output_num:
            break
        print("#{} {}".format(idx+1, k))
        for ingr in v.per_base_mat:
            print("{} {:.5g}".format(ingr, v.per_base_mat[ingr]))
        print("electric {:.5g} {:.5g}".format(v.electric, v.per_electric))
        print("time {:.5g}".format(v.per_time))
        print("money {:.5g} {:.5g}".format(v.money_cost, v.per_money))

if __name__ == '__main__':
    read_table("ResonancePrison.xlsx")
    enable_purchase = click.prompt("是否检索原料采购（y/n）", default='y') == 'y'
    calculate(enable_purchase)
    output()